import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ── Config ────────────────────────────────────────────────────────────────────
# All paths are relative to the scripts/ folder
BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_PATH   = os.path.join(BASE_DIR, 'data', 'processed', 'accepted_features.csv')
OUTPUT_DIR  = os.path.join(BASE_DIR, 'outputs')

# Report filename includes today's date so each run produces a new file
# e.g. credit_portfolio_report_2026-05-19.xlsx
today       = datetime.today().strftime('%Y-%m-%d')
OUTPUT_PATH = os.path.join(OUTPUT_DIR, f'credit_portfolio_report_{today}.xlsx')

print(f"Data path : {DATA_PATH}")
print(f"Output    : {OUTPUT_PATH}")

# ── Load Data ─────────────────────────────────────────────────────────────────
print("\nLoading data...")
df = pd.read_csv(DATA_PATH, low_memory=False)
df['year'] = df['vintage'].str[:4].astype(int)
print(f"Loaded {len(df):,} rows")
# ── Calculate Metrics ─────────────────────────────────────────────────────────
print("Calculating metrics...")

# 1. Portfolio summary — single numbers for the cover sheet
summary = {
    'Report Date'               : today,
    'Total Loans'               : f"{len(df):,}",
    'Total Funded Amount (USD)' : f"{df['funded_amnt'].sum():,.0f}",
    'Overall Default Rate'      : f"{df['is_default'].mean()*100:.2f}%",
    'Average Interest Rate'     : f"{df['int_rate'].mean():.2f}%",
    'Average Loan Amount (USD)' : f"{df['loan_amnt'].mean():,.0f}",
    'Average DTI'               : f"{df['dti'].mean():.2f}%",
}

# 2. Default rate by grade
grade_summary = (
    df.groupby('grade')
    .agg(
        Total_Loans   = ('loan_amnt', 'count'),
        Default_Rate  = ('is_default', 'mean'),
        Avg_Int_Rate  = ('int_rate', 'mean'),
        Avg_Loan_Amnt = ('loan_amnt', 'mean')
    )
    .reindex(['A','B','C','D','E','F','G'])
    .reset_index()
)
grade_summary['Default_Rate'] = (grade_summary['Default_Rate'] * 100).round(2)
grade_summary['Avg_Int_Rate']  = grade_summary['Avg_Int_Rate'].round(2)
grade_summary['Avg_Loan_Amnt'] = grade_summary['Avg_Loan_Amnt'].round(0)

# 3. Yearly trend
yearly = (
    df.groupby('year')
    .agg(
        Total_Loans   = ('loan_amnt', 'count'),
        Default_Rate  = ('is_default', 'mean'),
        Avg_Int_Rate  = ('int_rate', 'mean'),
        Avg_Loan_Amnt = ('loan_amnt', 'mean')
    )
    .reset_index()
)
yearly['Default_Rate'] = (yearly['Default_Rate'] * 100).round(2)
yearly['Avg_Int_Rate']  = yearly['Avg_Int_Rate'].round(2)
yearly['Avg_Loan_Amnt'] = yearly['Avg_Loan_Amnt'].round(0)

# 4. Default rate by purpose
purpose_summary = (
    df.groupby('purpose')
    .agg(
        Total_Loans  = ('loan_amnt', 'count'),
        Default_Rate = ('is_default', 'mean'),
        Avg_Int_Rate = ('int_rate', 'mean')
    )
    .sort_values('Default_Rate', ascending=False)
    .reset_index()
)
purpose_summary['Default_Rate'] = (purpose_summary['Default_Rate'] * 100).round(2)
purpose_summary['Avg_Int_Rate']  = purpose_summary['Avg_Int_Rate'].round(2)

print("Metrics calculated.")

# ── Write Excel Report ────────────────────────────────────────────────────────
print("\nWriting Excel report...")

wb = openpyxl.Workbook()

# openpyxl creates a default sheet called 'Sheet' when you make a new workbook
# We don't need it so we remove it
wb.remove(wb.active)

# ── Define reusable styles ────────────────────────────────────────────────────
# Header style — dark blue background, white bold text
header_font    = Font(bold=True, color='FFFFFF', size=11)
header_fill    = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_align   = Alignment(horizontal='center', vertical='center')

# Title style — larger bold text for sheet titles
title_font     = Font(bold=True, size=13)

# Subheader style — medium blue, white text
sub_font       = Font(bold=True, color='FFFFFF', size=10)
sub_fill       = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')

def style_header_row(ws, row_num, num_cols):
    """Apply dark blue header styling to a specific row."""
    for col in range(1, num_cols + 1):
        cell            = ws.cell(row=row_num, column=col)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = header_align

def auto_col_width(ws):
    """Auto-fit column widths based on content length."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
ws_summary = wb.create_sheet('Portfolio Summary')

# Sheet title
ws_summary['A1'] = 'Credit Portfolio Report — Lending Club 2007–2018'
ws_summary['A1'].font = title_font

ws_summary['A2'] = f'Generated on: {today}'
ws_summary['A2'].font = Font(italic=True, size=10, color='888888')

# Leave a blank row then write the summary metrics
ws_summary['A4'] = 'Metric'
ws_summary['B4'] = 'Value'
style_header_row(ws_summary, 4, 2)

# Write each metric as a row
row = 5
for metric, value in summary.items():
    ws_summary.cell(row=row, column=1).value = metric
    ws_summary.cell(row=row, column=2).value = value
    # Alternate row shading for readability
    if row % 2 == 0:
        light_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        ws_summary.cell(row=row, column=1).fill = light_fill
        ws_summary.cell(row=row, column=2).fill = light_fill
    row += 1

auto_col_width(ws_summary)

# ── Sheet 2: Grade Analysis ───────────────────────────────────────────────────
ws_grade = wb.create_sheet('Grade Analysis')

ws_grade['A1'] = 'Default Rate by Loan Grade'
ws_grade['A1'].font = title_font

# Write column headers
headers = ['Grade', 'Total Loans', 'Default Rate (%)', 'Avg Interest Rate (%)', 'Avg Loan Amount (USD)']
for col, header in enumerate(headers, start=1):
    ws_grade.cell(row=3, column=col).value = header
style_header_row(ws_grade, 3, len(headers))

# Write data rows
for row, record in enumerate(grade_summary.itertuples(), start=4):
    ws_grade.cell(row=row, column=1).value = record.grade
    ws_grade.cell(row=row, column=2).value = record.Total_Loans
    ws_grade.cell(row=row, column=3).value = record.Default_Rate
    ws_grade.cell(row=row, column=4).value = record.Avg_Int_Rate
    ws_grade.cell(row=row, column=5).value = record.Avg_Loan_Amnt
    if row % 2 == 0:
        light_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        for col in range(1, 6):
            ws_grade.cell(row=row, column=col).fill = light_fill

auto_col_width(ws_grade)

# ── Sheet 3: Yearly Trends ────────────────────────────────────────────────────
ws_yearly = wb.create_sheet('Yearly Trends')

ws_yearly['A1'] = 'Portfolio Performance by Year'
ws_yearly['A1'].font = title_font

headers = ['Year', 'Total Loans', 'Default Rate (%)', 'Avg Interest Rate (%)', 'Avg Loan Amount (USD)']
for col, header in enumerate(headers, start=1):
    ws_yearly.cell(row=3, column=col).value = header
style_header_row(ws_yearly, 3, len(headers))

for row, record in enumerate(yearly.itertuples(), start=4):
    ws_yearly.cell(row=row, column=1).value = record.year
    ws_yearly.cell(row=row, column=2).value = record.Total_Loans
    ws_yearly.cell(row=row, column=3).value = record.Default_Rate
    ws_yearly.cell(row=row, column=4).value = record.Avg_Int_Rate
    ws_yearly.cell(row=row, column=5).value = record.Avg_Loan_Amnt
    if row % 2 == 0:
        light_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        for col in range(1, 6):
            ws_yearly.cell(row=row, column=col).fill = light_fill

auto_col_width(ws_yearly)

# ── Sheet 4: Purpose Analysis ─────────────────────────────────────────────────
ws_purpose = wb.create_sheet('Purpose Analysis')

ws_purpose['A1'] = 'Default Rate by Loan Purpose'
ws_purpose['A1'].font = title_font

headers = ['Purpose', 'Total Loans', 'Default Rate (%)', 'Avg Interest Rate (%)']
for col, header in enumerate(headers, start=1):
    ws_purpose.cell(row=3, column=col).value = header
style_header_row(ws_purpose, 3, len(headers))

for row, record in enumerate(purpose_summary.itertuples(), start=4):
    ws_purpose.cell(row=row, column=1).value = record.purpose
    ws_purpose.cell(row=row, column=2).value = record.Total_Loans
    ws_purpose.cell(row=row, column=3).value = record.Default_Rate
    ws_purpose.cell(row=row, column=4).value = record.Avg_Int_Rate
    if row % 2 == 0:
        light_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        for col in range(1, 5):
            ws_purpose.cell(row=row, column=col).fill = light_fill

auto_col_width(ws_purpose)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"Report saved to {OUTPUT_PATH}")